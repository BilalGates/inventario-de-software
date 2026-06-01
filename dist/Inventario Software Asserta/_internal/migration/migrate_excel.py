from __future__ import annotations

import argparse
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy import text

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from database.connection import get_engine
from modules.software import generar_codigo_software
from utils.normalizer import clean_version, normalize_equipo_nombre, normalize_nombre


SHEET_TO_DEPT = {
    "Terminales_Gerencia": "gerencia",
    "Terminales_Administrativos": "administracion",
    "Terminales_IT": "it",
    "Terminales_Analytics": "data_science",
    "Terminales_Silicon": "silicon",
    "Servidores": "servidores",
}

HEADER_ALIASES = {
    "codigo": ["id software"],
    "nombre": ["nombre del software", "nombre"],
    "fabricante": ["fabricante proveedor", "fabricante / proveedor", "editor", "fabricante"],
    "version": ["version actual", "version"],
    "clasificacion": ["clasificacion informacion tratada"],
    "dispositivos": ["dispositivos"],
    "fecha_actualizacion": ["fecha ultima actualizacion", "fecha ultima revision"],
    "en_guia": ["en guia 105", "en guia"],
    "obs_elena": ["observaciones elena"],
    "obs_toni": ["observaciones toni"],
    "usuario": ["usuario"],
}

class WarningLog:
    def __init__(self) -> None:
        self.items: list[str] = []

    def add(self, message: str) -> None:
        self.items.append(message)

    def write(self, path: Path) -> None:
        path.write_text("\n".join(self.items), encoding="utf-8")


def _norm_header(value: Any) -> str:
    raw = str(value or "").strip().lower()
    raw = unicodedata.normalize("NFKD", raw)
    raw = "".join(ch for ch in raw if not unicodedata.combining(ch))
    for ch in "¿?/:()[]":
        raw = raw.replace(ch, " ")
    return " ".join(raw.split())


def _header_index(row: tuple[Any, ...]) -> dict[str, int]:
    normalized = [_norm_header(value) for value in row]
    mapping: dict[str, int] = {}
    for key, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            alias_norm = _norm_header(alias)
            if alias_norm in normalized:
                mapping[key] = normalized.index(alias_norm)
                break
    return mapping


def _get(row: tuple[Any, ...], mapping: dict[str, int], key: str):
    idx = mapping.get(key)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _clean_text(value: Any) -> str | None:
    if value is None:
        return None
    val = str(value).strip()
    return val if val not in ("", "-", "–", "nan", "None") else None


def _parse_date(value: Any):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    val = str(value).strip()
    if not val:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(val, fmt).date()
        except ValueError:
            continue
    return None


def _parse_bool(value: Any):
    if value is None:
        return None
    val = _norm_header(value)
    if val in {"si", "s", "yes", "true", "1"}:
        return True
    if val in {"no", "n", "false", "0"}:
        return False
    return None


def _clean_fabricante(value: Any, warnings: WarningLog, context: str) -> str | None:
    fabricante = _clean_text(value)
    if not fabricante:
        return None
    if fabricante == "${AppPublisher}":
        warnings.add(f"{context}: fabricante '${{AppPublisher}}' guardado como NULL")
        return None
    if "\\" in fabricante:
        warnings.add(f"{context}: fabricante con barra invertida '{fabricante}' guardado como NULL")
        return None
    return fabricante


def _get_dept_id(db, codigo: str) -> int:
    dept_id = db.execute(
        text("SELECT id FROM departamentos WHERE codigo = :codigo"),
        {"codigo": codigo},
    ).scalar()
    if not dept_id:
        raise RuntimeError(f"No existe el departamento '{codigo}'. Ejecute database/seed.sql primero.")
    return int(dept_id)


def _get_or_create_equipo(db, departamento_id: int, nombre: str, counters: dict) -> int:
    nombre_norm = normalize_equipo_nombre(nombre)
    existing = db.execute(
        text(
            """
            SELECT id
            FROM equipos
            WHERE departamento_id = :departamento_id
              AND nombre_norm = :nombre_norm
            """
        ),
        {"departamento_id": departamento_id, "nombre_norm": nombre_norm},
    ).scalar()
    if existing:
        return int(existing)
    result = db.execute(
        text(
            """
            INSERT INTO equipos (departamento_id, nombre, nombre_norm)
            VALUES (:departamento_id, :nombre, :nombre_norm)
            """
        ),
        {"departamento_id": departamento_id, "nombre": nombre.strip(), "nombre_norm": nombre_norm},
    )
    counters["equipos_creados"] += 1
    return int(result.lastrowid)


def _get_software_id(db, departamento_id: int, nombre_norm: str) -> int | None:
    software_id = db.execute(
        text(
            """
            SELECT id
            FROM software
            WHERE departamento_id = :departamento_id
              AND nombre_norm = :nombre_norm
            """
        ),
        {"departamento_id": departamento_id, "nombre_norm": nombre_norm},
    ).scalar()
    return int(software_id) if software_id else None


def _upsert_software(db, departamento_id: int, data: dict, counters: dict) -> int:
    nombre_norm = normalize_nombre(data["nombre"])
    software_id = _get_software_id(db, departamento_id, nombre_norm)
    codigo = data.get("codigo") or (generar_codigo_software(db, departamento_id) if not software_id else None)
    params = {**data, "departamento_id": departamento_id, "nombre_norm": nombre_norm, "codigo": codigo}
    if software_id:
        params["software_id"] = software_id
        db.execute(
            text(
                """
                UPDATE software
                SET codigo = COALESCE(:codigo, codigo),
                    nombre = :nombre,
                    fabricante = :fabricante,
                    version_referencia = :version_referencia,
                    clasificacion_informacion = :clasificacion_informacion,
                    en_guia_105 = :en_guia_105,
                    observaciones_elena = :observaciones_elena,
                    observaciones_toni = :observaciones_toni,
                    fecha_ultima_actualizacion = :fecha_ultima_actualizacion,
                    activo = TRUE
                WHERE id = :software_id
                """
            ),
            params,
        )
        return software_id

    result = db.execute(
        text(
            """
            INSERT INTO software (
                departamento_id, codigo, nombre, nombre_norm, fabricante,
                version_referencia, clasificacion_informacion, en_guia_105,
                observaciones_elena, observaciones_toni, fecha_ultima_actualizacion
            )
            VALUES (
                :departamento_id, :codigo, :nombre, :nombre_norm, :fabricante,
                :version_referencia, :clasificacion_informacion, :en_guia_105,
                :observaciones_elena, :observaciones_toni, :fecha_ultima_actualizacion
            )
            """
        ),
        params,
    )
    counters["software_catalogo"] += 1
    return int(result.lastrowid)


def _link_software_equipo(db, software_id: int, equipo_id: int, version: str | None, revision_date, counters: dict) -> None:
    db.execute(
        text(
            """
            INSERT INTO software_equipo (
                software_id, equipo_id, version_detectada, fecha_ultima_deteccion, presente
            )
            VALUES (:software_id, :equipo_id, :version_detectada, :fecha_ultima_deteccion, TRUE)
            ON DUPLICATE KEY UPDATE
                version_detectada = VALUES(version_detectada),
                fecha_ultima_deteccion = VALUES(fecha_ultima_deteccion),
                presente = TRUE
            """
        ),
        {
            "software_id": software_id,
            "equipo_id": equipo_id,
            "version_detectada": version,
            "fecha_ultima_deteccion": revision_date or date.today(),
        },
    )
    counters["software_equipo"] += 1


def _iter_data_rows(sheet):
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {}, []
    mapping = _header_index(rows[0])
    return mapping, rows[1:]


def migrate_department_sheet(db, sheet, departamento_codigo: str, counters: dict, warnings: WarningLog) -> None:
    departamento_id = _get_dept_id(db, departamento_codigo)
    mapping, rows = _iter_data_rows(sheet)
    for row_number, row in enumerate(rows, start=2):
        nombre = _clean_text(_get(row, mapping, "nombre"))
        if not nombre:
            continue
        context = f"{sheet.title} fila {row_number} ({nombre})"
        version = clean_version(_get(row, mapping, "version"))
        raw_version = _get(row, mapping, "version")
        if isinstance(raw_version, (int, float)):
            warnings.add(f"{context}: version numerica '{raw_version}' guardada como texto '{version}'")
        update_date = _parse_date(_get(row, mapping, "fecha_actualizacion"))
        software_id = _upsert_software(
            db,
            departamento_id,
            {
                "codigo": _clean_text(_get(row, mapping, "codigo")),
                "nombre": nombre,
                "fabricante": _clean_fabricante(_get(row, mapping, "fabricante"), warnings, context),
                "version_referencia": version,
                "clasificacion_informacion": _clean_text(_get(row, mapping, "clasificacion")) or "Media",
                "en_guia_105": _parse_bool(_get(row, mapping, "en_guia")),
                "observaciones_elena": _clean_text(_get(row, mapping, "obs_elena")),
                "observaciones_toni": _clean_text(_get(row, mapping, "obs_toni")),
                "fecha_ultima_actualizacion": update_date,
            },
            counters,
        )
        dispositivos = _clean_text(_get(row, mapping, "dispositivos"))
        if not dispositivos:
            continue
        for equipo_nombre in [part.strip() for part in dispositivos.split(",") if part.strip()]:
            equipo_id = _get_or_create_equipo(db, departamento_id, equipo_nombre, counters)
            _link_software_equipo(db, software_id, equipo_id, version, update_date, counters)


def _find_equipo_global(db, usuario: str) -> int | None:
    nombre_norm = normalize_equipo_nombre(usuario)
    equipo_id = db.execute(
        text("SELECT id FROM equipos WHERE nombre_norm = :nombre_norm LIMIT 1"),
        {"nombre_norm": nombre_norm},
    ).scalar()
    return int(equipo_id) if equipo_id else None


def migrate_autorizado_sheet(db, sheet, counters: dict, warnings: WarningLog) -> None:
    mapping, rows = _iter_data_rows(sheet)
    for row_number, row in enumerate(rows, start=2):
        nombre = _clean_text(_get(row, mapping, "nombre"))
        if not nombre:
            continue
        context = f"{sheet.title} fila {row_number} ({nombre})"
        usuario = _clean_text(_get(row, mapping, "usuario"))
        equipo_id = _find_equipo_global(db, usuario) if usuario else None
        usuario_texto = None if equipo_id else usuario
        version = clean_version(_get(row, mapping, "version"))
        if isinstance(_get(row, mapping, "version"), (int, float)):
            warnings.add(f"{context}: version numerica guardada como texto '{version}'")
        db.execute(
            text(
                """
                INSERT INTO software_autorizado (
                    nombre, fabricante, tipo, version, equipo_id, usuario_texto
                )
                VALUES (
                    :nombre, :fabricante, :tipo, :version, :equipo_id, :usuario_texto
                )
                """
            ),
            {
                "nombre": nombre,
                "fabricante": _clean_fabricante(_get(row, mapping, "fabricante"), warnings, context),
                "tipo": _clean_text(_get(row, mapping, "tipo")),
                "version": version,
                "equipo_id": equipo_id,
                "usuario_texto": usuario_texto,
            },
        )
        counters["software_autorizado"] += 1


def migrate(path: Path) -> dict:
    warnings = WarningLog()
    counters = {
        "departamentos": 0,
        "equipos_creados": 0,
        "software_catalogo": 0,
        "software_equipo": 0,
        "software_autorizado": 0,
    }
    workbook = openpyxl.load_workbook(path, data_only=True)
    engine = get_engine()
    with engine.begin() as db:
        counters["departamentos"] = db.execute(text("SELECT COUNT(*) FROM departamentos")).scalar() or 0
        for sheet_name, dept_codigo in SHEET_TO_DEPT.items():
            if sheet_name not in workbook.sheetnames:
                warnings.add(f"Hoja no encontrada: {sheet_name}")
                continue
            migrate_department_sheet(db, workbook[sheet_name], dept_codigo, counters, warnings)
        if "SW_Autorizado" in workbook.sheetnames:
            migrate_autorizado_sheet(db, workbook["SW_Autorizado"], counters, warnings)
        else:
            warnings.add("Hoja no encontrada: SW_Autorizado")

    warnings.write(ROOT / "migration" / "migrate_warnings.log")
    counters["warnings"] = len(warnings.items)
    return counters


def main() -> int:
    parser = argparse.ArgumentParser(description="Migra el Excel historico de inventario a MySQL.")
    parser.add_argument("--file", required=True, help="Ruta al Inventario_Software_ENS_Por_Departamento.xlsx")
    args = parser.parse_args()
    path = Path(args.file).expanduser().resolve()
    if not path.exists():
        print(f"No existe el fichero: {path}", file=sys.stderr)
        return 1

    counters = migrate(path)
    print("=== Migración completada ===")
    print(f"Departamentos: {counters['departamentos']}")
    print(f"Equipos creados: {counters['equipos_creados']}")
    print(f"Software (catálogo): {counters['software_catalogo']} entradas")
    print(f"Software_equipo (links): {counters['software_equipo']}")
    print(f"Software_autorizado: {counters['software_autorizado']}")
    print(f"Warnings: {counters['warnings']} (ver migration/migrate_warnings.log)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
