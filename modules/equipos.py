from __future__ import annotations

from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import text

from utils.normalizer import normalize_equipo_nombre


def _safe_sheet_name(name: str) -> str:
    invalid = "[]:*?/\\"
    cleaned = "".join("_" if char in invalid else char for char in name)
    return cleaned[:31] or "Equipos"


def listar_equipos(
    db,
    departamento_id: int | None = None,
    solo_activos: bool = True,
    es_servidor: bool | None = None,
) -> list[dict]:
    sql = """
        SELECT
            e.id,
            e.departamento_id,
            e.nombre,
            e.nombre_norm,
            e.activo,
            e.es_servidor,
            e.fecha_alta,
            e.fecha_baja,
            e.notas,
            e.tipo_dispositivo,
            e.marca_modelo,
            e.num_serie,
            e.mac_address,
            e.sistema_operativo,
            e.procesador,
            e.ram,
            e.almacenamiento,
            e.responsable,
            e.ubicacion,
            e.coste,
            e.fecha_adquisicion,
            imp.fecha_importacion AS ultima_importacion,
            imp.metodo AS ultimo_metodo,
            imp.n_total AS ultimo_total,
            imp.n_nuevos AS ultimo_nuevos,
            imp.n_actualizados AS ultimo_actualizados,
            imp.n_eliminados AS ultimo_eliminados,
            imp.n_cambios_version AS ultimo_cambios_version,
            COALESCE(sw.total_software, 0) AS total_software_activo,
            d.nombre AS departamento_nombre
        FROM equipos e
        JOIN departamentos d ON d.id = e.departamento_id
        LEFT JOIN (
            SELECT i.*
            FROM importaciones i
            JOIN (
                SELECT equipo_id, MAX(fecha_importacion) AS max_fecha
                FROM importaciones
                WHERE confirmada = TRUE
                GROUP BY equipo_id
            ) latest
              ON latest.equipo_id = i.equipo_id
             AND latest.max_fecha = i.fecha_importacion
            WHERE i.confirmada = TRUE
        ) imp ON imp.equipo_id = e.id
        LEFT JOIN (
            SELECT equipo_id, COUNT(*) AS total_software
            FROM software_equipo
            WHERE presente = TRUE
            GROUP BY equipo_id
        ) sw ON sw.equipo_id = e.id
        WHERE 1 = 1
    """
    params = {}
    if departamento_id is not None:
        sql += " AND e.departamento_id = :departamento_id"
        params["departamento_id"] = departamento_id
    if solo_activos:
        sql += " AND e.activo = TRUE"
    if es_servidor is not None:
        sql += " AND e.es_servidor = :es_servidor"
        params["es_servidor"] = es_servidor
    sql += " ORDER BY e.activo DESC, d.nombre ASC, e.nombre ASC"
    return [dict(row) for row in db.execute(text(sql), params).mappings().all()]


def _as_date(value) -> date | None:
    if value is None:
        return None
    try:
        if value != value:
            return None
    except TypeError:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if hasattr(value, "date"):
        converted = value.date()
        try:
            if converted != converted:
                return None
        except TypeError:
            return None
        return converted if isinstance(converted, date) else None
    return None


def estado_importacion(fecha_importacion) -> str:
    fecha = _as_date(fecha_importacion)
    if not fecha:
        return "🔴 Nunca importado"
    days = (date.today() - fecha).days
    if days < 30:
        return f"🟢 Hace {days} días"
    if days <= 60:
        return f"🟡 Hace {days} días"
    return f"🔴 Hace {days} días"


def alertas_equipo(equipo: dict) -> list[str]:
    alerts = []
    estado = estado_importacion(equipo.get("ultima_importacion"))
    if estado.startswith("🟡") or estado.startswith("🔴"):
        alerts.append("Importación pendiente o nunca realizada.")
    if not equipo.get("responsable"):
        alerts.append("Sin responsable asignado.")
    missing_hardware = [
        field
        for field in ("procesador", "ram", "sistema_operativo")
        if not equipo.get(field)
    ]
    if missing_hardware:
        alerts.append(f"Campos hardware vacíos: {', '.join(missing_hardware)}.")
    return alerts


def listar_estado_importaciones(db, departamento_id: int | None = None) -> list[dict]:
    equipos = listar_equipos(db, departamento_id=departamento_id, solo_activos=True)
    rows = []
    for equipo in equipos:
        fecha = _as_date(equipo.get("ultima_importacion"))
        dias = None
        if fecha:
            dias = (date.today() - fecha).days
        rows.append(
            {
                **equipo,
                "dias_desde_importacion": dias,
                "estado_importacion": estado_importacion(fecha),
            }
        )
    return rows


def obtener_equipo(db, equipo_id: int) -> dict | None:
    row = db.execute(
        text(
            """
            SELECT e.*, d.nombre AS departamento_nombre, d.prefijo_id
            FROM equipos e
            JOIN departamentos d ON d.id = e.departamento_id
            WHERE e.id = :equipo_id
            """
        ),
        {"equipo_id": equipo_id},
    ).mappings().first()
    return dict(row) if row else None


def existe_equipo(db, departamento_id: int, nombre: str) -> bool:
    nombre_norm = normalize_equipo_nombre(nombre)
    row = db.execute(
        text(
            """
            SELECT id
            FROM equipos
            WHERE departamento_id = :departamento_id
              AND nombre_norm = :nombre_norm
            LIMIT 1
            """
        ),
        {"departamento_id": departamento_id, "nombre_norm": nombre_norm},
    ).first()
    return row is not None


def crear_equipo(
    db,
    departamento_id: int,
    nombre: str,
    notas: str | None = None,
    es_servidor: bool = False,
) -> int:
    nombre_norm = normalize_equipo_nombre(nombre)
    result = db.execute(
        text(
            """
            INSERT INTO equipos (departamento_id, nombre, nombre_norm, notas, es_servidor)
            VALUES (:departamento_id, :nombre, :nombre_norm, :notas, :es_servidor)
            """
        ),
        {
            "departamento_id": departamento_id,
            "nombre": nombre.strip(),
            "nombre_norm": nombre_norm,
            "notas": notas or None,
            "es_servidor": es_servidor,
        },
    )
    return int(result.lastrowid)


def exportar_equipos_excel(equipos: list[dict], titulo: str = "Equipos") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_name(titulo)
    columns = [
        ("id", "ID"),
        ("departamento_id", "Departamento ID"),
        ("departamento_nombre", "Departamento"),
        ("nombre", "Equipo"),
        ("nombre_norm", "Nombre normalizado"),
        ("notas", "Usuario"),
        ("activo", "Activo"),
        ("tipo_dispositivo", "Tipo dispositivo"),
        ("marca_modelo", "Marca / Modelo"),
        ("num_serie", "Nº de serie"),
        ("mac_address", "Dirección MAC"),
        ("sistema_operativo", "Sistema operativo"),
        ("procesador", "Procesador"),
        ("ram", "RAM"),
        ("almacenamiento", "Almacenamiento"),
        ("responsable", "Responsable"),
        ("ubicacion", "Ubicación"),
        ("coste", "Coste"),
        ("fecha_adquisicion", "Fecha adquisición"),
        ("ultima_importacion", "Última importación"),
        ("total_software_activo", "Software activo instalado"),
        ("fecha_alta", "Fecha alta"),
        ("fecha_baja", "Fecha baja"),
    ]
    ws.append([label for _, label in columns])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="366092")
    for equipo in equipos:
        ws.append([
            "Sí" if key == "activo" and equipo.get(key) else "No" if key == "activo" else equipo.get(key)
            for key, _ in columns
        ])
    for column in ws.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 60)
        ws.column_dimensions[column[0].column_letter].width = width
    ws.freeze_panes = "A2"
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def dar_baja_equipo(db, equipo_id: int) -> None:
    db.execute(
        text(
            """
            UPDATE equipos
            SET activo = FALSE, fecha_baja = :fecha_baja
            WHERE id = :equipo_id
            """
        ),
        {"equipo_id": equipo_id, "fecha_baja": date.today()},
    )


def actualizar_usuario_dispositivo(db, equipo_id: int, usuario: str | None) -> None:
    db.execute(
        text(
            """
            UPDATE equipos
            SET notas = :usuario
            WHERE id = :equipo_id
            """
        ),
        {"equipo_id": equipo_id, "usuario": usuario or None},
    )


def _parse_cost(value) -> float | None:
    if value is None:
        return None
    try:
        cleaned = str(value).replace(",", ".").replace(" ", "")
        return float(cleaned)
    except (ValueError, TypeError):
        return None


EQUIPO_UPDATE_FIELDS = [
    "tipo_dispositivo", "marca_modelo", "num_serie", "mac_address",
    "sistema_operativo", "procesador", "ram", "almacenamiento",
    "responsable", "ubicacion", "notas",
]


def _build_update_sql(equipo_id: int, data: dict) -> tuple[str, dict]:
    sets = []
    params: dict = {"equipo_id": equipo_id}
    for field in EQUIPO_UPDATE_FIELDS:
        if field in data:
            sets.append(f"{field} = :{field}")
            params[field] = data.get(field) or None
    if "activo" in data:
        sets.append("activo = :activo")
        params["activo"] = bool(data["activo"])
    if data.get("coste") is not None:
        coste = _parse_cost(data["coste"])
        if coste is not None:
            sets.append("coste = :coste")
            params["coste"] = coste
    if data.get("fecha_adquisicion") is not None:
        sets.append("fecha_adquisicion = :fecha_adquisicion")
        params["fecha_adquisicion"] = data["fecha_adquisicion"]
    if not sets:
        return "", params
    sql = f"UPDATE equipos SET {', '.join(sets)} WHERE id = :equipo_id"
    return sql, params


def importar_equipos_desde_lista(
    db,
    equipos_data: list[dict],
    departamentos: list[dict],
    es_servidor: bool = False,
) -> dict:
    dept_por_nombre: dict[str, int] = {}
    for d in departamentos:
        dept_por_nombre[d["nombre"].strip().lower()] = d["id"]

    resultado = {"creados": 0, "actualizados": 0, "omitidos": 0, "errores": []}

    for item in equipos_data:
        nombre = (item.get("nombre") or "").strip()
        if not nombre:
            resultado["omitidos"] += 1
            continue

        dept_name = (item.get("departamento_nombre") or "").strip().lower()
        departamento_id = dept_por_nombre.get(dept_name)
        if departamento_id is None:
            resultado["errores"].append(f"Departamento '{item.get('departamento_nombre')}' no encontrado para '{nombre}'")
            continue

        nombre_norm = normalize_equipo_nombre(nombre)
        existente = db.execute(
            text(
                """
                SELECT id, notas, activo
                FROM equipos
                WHERE departamento_id = :dept_id AND nombre_norm = :norm
                LIMIT 1
                """
            ),
            {"dept_id": departamento_id, "norm": nombre_norm},
        ).mappings().first()

        if existente:
            sql, params = _build_update_sql(existente["id"], item)
            if sql:
                db.execute(text(sql), params)
            resultado["actualizados"] += 1
        else:
            notas = item.get("notas") or None
            activo = item.get("activo")
            if activo is None:
                activo = True
            result = db.execute(
                text(
                    """
                    INSERT INTO equipos
                        (departamento_id, nombre, nombre_norm, notas, activo, es_servidor)
                    VALUES
                        (:dept_id, :nombre, :norm, :notas, :activo, :es_servidor)
                    """
                ),
                {
                    "dept_id": departamento_id,
                    "nombre": nombre,
                    "norm": nombre_norm,
                    "notas": notas,
                    "activo": bool(activo),
                    "es_servidor": es_servidor,
                },
            )
            equipo_id = int(result.lastrowid)
            sql, params = _build_update_sql(equipo_id, item)
            if sql:
                db.execute(text(sql), params)
            resultado["creados"] += 1

    return resultado
