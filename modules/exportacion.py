from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import text

from modules.software import listar_inventario


EXPORT_COLUMNS = [
    ("codigo", "ID Software"),
    ("nombre", "Nombre del Software"),
    ("fabricante", "Fabricante / Proveedor"),
    ("version_referencia", "Versión"),
    ("dispositivos", "Dispositivos"),
    ("fecha_ultima_actualizacion", "Última actualización"),
    ("clasificacion_informacion", "Clasificación"),
    ("en_guia_105", "En Guia 105"),
    ("observaciones", "Observaciones"),
]


def _cell_value(value):
    if isinstance(value, bool):
        return "Sí" if value else "No"
    return value


def _safe_sheet_name(name: str) -> str:
    invalid = "[]:*?/\\"
    cleaned = "".join("_" if char in invalid else char for char in name)
    return cleaned[:31] or "Hoja"


def _write_rows(ws, headers: list[str], rows: list[list]):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="366092")
    for row in rows:
        ws.append(row)
    for column in ws.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 60)
        ws.column_dimensions[column[0].column_letter].width = width
    ws.freeze_panes = "A2"


def generar_excel(
    db,
    departamento_ids: list[int],
    detalle_por_equipo: bool = False,
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    for departamento_id in departamento_ids:
        dept = db.execute(
            text("SELECT id, nombre FROM departamentos WHERE id = :id"),
            {"id": departamento_id},
        ).mappings().first()
        if not dept:
            continue
        inventario = listar_inventario(db, departamento_id)
        ws = wb.create_sheet(_safe_sheet_name(dept["nombre"]))
        headers = [label for _, label in EXPORT_COLUMNS]
        rows = [[_cell_value(item.get(key)) for key, _ in EXPORT_COLUMNS] for item in inventario]
        _write_rows(ws, headers, rows)

        if detalle_por_equipo:
            equipos = db.execute(
                text(
                    """
                    SELECT id, nombre
                    FROM equipos
                    WHERE departamento_id = :departamento_id
                      AND activo = TRUE
                    ORDER BY nombre
                    """
                ),
                {"departamento_id": departamento_id},
            ).mappings().all()
            for equipo in equipos:
                detail_rows = db.execute(
                    text(
                        """
                        SELECT
                            s.codigo,
                            s.nombre,
                            s.fabricante,
                            swe.version_detectada,
                            swe.fecha_instalacion,
                            swe.tamano,
                            swe.fecha_ultima_deteccion
                        FROM software_equipo swe
                        JOIN software s ON s.id = swe.software_id
                        WHERE swe.equipo_id = :equipo_id
                          AND swe.presente = TRUE
                          AND s.activo = TRUE
                        ORDER BY s.nombre
                        """
                    ),
                    {"equipo_id": equipo["id"]},
                ).mappings().all()
                ws_detail = wb.create_sheet(_safe_sheet_name(f"{dept['nombre']}-{equipo['nombre']}"))
                _write_rows(
                    ws_detail,
                    [
                        "ID Software",
                        "Nombre del Software",
                        "Fabricante / Proveedor",
                        "Versión Detectada",
                        "Fecha Instalación",
                        "Tamaño",
                        "Fecha Última Detección",
                    ],
                    [[_cell_value(value) for value in row.values()] for row in detail_rows],
                )

    if not wb.sheetnames:
        ws = wb.create_sheet("Inventario")
        _write_rows(ws, ["Sin datos"], [])

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def generar_excel_importaciones(db, departamento_id: int) -> bytes:
    rows = db.execute(
        text(
            """
            SELECT
                e.nombre AS equipo,
                i.fecha_importacion,
                i.metodo,
                i.n_total,
                i.n_nuevos,
                i.n_actualizados,
                i.n_eliminados,
                i.n_cambios_version
            FROM importaciones i
            JOIN equipos e ON e.id = i.equipo_id
            WHERE e.departamento_id = :departamento_id
              AND i.confirmada = TRUE
            ORDER BY i.fecha_importacion DESC, i.id DESC
            """
        ),
        {"departamento_id": departamento_id},
    ).mappings().all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Importaciones"
    headers = [
        "Equipo",
        "Fecha",
        "Método",
        "Total software",
        "Nuevos",
        "Actualizados",
        "Eliminados",
        "Cambios de versión",
    ]
    _write_rows(
        ws,
        headers,
        [
            [
                row["equipo"],
                row["fecha_importacion"],
                row["metodo"],
                row["n_total"],
                row["n_nuevos"],
                row["n_actualizados"],
                row["n_eliminados"],
                row["n_cambios_version"],
            ]
            for row in rows
        ],
    )
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def generar_excel_completo(db) -> bytes:
    from modules.autorizado import listar_autorizado_agrupado
    from modules.equipos import listar_equipos
    from modules.software import listar_departamentos

    wb = Workbook()
    wb.remove(wb.active)

    departamentos = listar_departamentos(db)

    inventario_headers = [
        "ID", "Programa", "Fabricante", "Versión", "Dispositivos",
        "Última actualización", "Clasificación", "Guía 105", "Observaciones",
    ]

    for dept in departamentos:
        inventario = listar_inventario(db, dept["id"])
        ws = wb.create_sheet(_safe_sheet_name(dept["nombre"]))
        rows = [
            [_cell_value(item.get(key)) for key, _ in EXPORT_COLUMNS]
            for item in inventario
        ]
        _write_rows(ws, inventario_headers, rows)

    device_columns_gral = [
        ("id", "ID"),
        ("departamento_nombre", "Departamento"),
        ("nombre", "Dispositivo"),
        ("es_servidor", "¿Servidor?"),
        ("notas", "Usuario"),
        ("tipo_dispositivo", "Tipo dispositivo"),
        ("marca_modelo", "Marca / Modelo"),
        ("num_serie", "Nº de serie"),
        ("mac_address", "Dirección MAC"),
        ("sistema_operativo", "Sistema operativo"),
        ("procesador", "Procesador"),
        ("ram", "RAM"),
        ("almacenamiento", "Almacenamiento"),
        ("responsable", "Responsable"),
        ("ubicacion", "Ubicación"),
        ("coste", "Coste"),
        ("fecha_adquisicion", "Fecha adquisición"),
        ("ultima_importacion", "Última importación"),
        ("total_software_activo", "Software activo instalado"),
    ]
    all_devices = listar_equipos(db, es_servidor=False)
    if all_devices:
        ws_devices = wb.create_sheet("Dispositivos")
        _write_rows(
            ws_devices,
            [label for _, label in device_columns_gral],
            [[_cell_value(e.get(key)) for key, _ in device_columns_gral] for e in all_devices],
        )

    servidor_columns = [
        ("id", "Nº"),
        ("nombre", "Nombre del Equipo"),
        ("tipo_dispositivo", "Tipo de Dispositivo"),
        ("sistema_operativo", "Sistema Operativo"),
        ("procesador", "Procesador"),
        ("ram", "Memoria RAM"),
        ("almacenamiento", "Almacenamiento"),
        ("departamento_nombre", "Departamento"),
        ("ubicacion", "Ubicación Física"),
        ("notas", "Observaciones"),
        ("coste", "Coste"),
        ("fecha_adquisicion", "Fecha de Adquisición"),
    ]
    servidor_equipos = listar_equipos(db, es_servidor=True)
    if servidor_equipos:
        ws_serv = wb.create_sheet("Servidores")
        servidor_rows = []
        for e in servidor_equipos:
            row = [_cell_value(e.get(key)) for key, _ in servidor_columns]
            estado = "Activo" if e.get("activo") else "Inactivo"
            row.insert(9, estado)
            servidor_rows.append(row)
        servidor_headers = [label for _, label in servidor_columns]
        servidor_headers.insert(9, "Estado")
        _write_rows(ws_serv, servidor_headers, servidor_rows)

    autorizados = listar_autorizado_agrupado(db)
    ws_auth = wb.create_sheet("Software Autorizado")
    _write_rows(
        ws_auth,
        [
            "Programa",
            "Fabricantes",
            "Versiones disponibles",
            "Departamentos",
            "Equipos/Usuarios",
            "Observaciones",
            "Última autorización",
            "Registros",
        ],
        [
            [
                row.get("nombre"),
                row.get("fabricantes"),
                row.get("versiones"),
                row.get("departamentos"),
                row.get("equipos_usuarios"),
                row.get("observaciones"),
                row.get("fecha_reciente"),
                row.get("registros"),
            ]
            for row in autorizados
        ],
    )

    if not wb.sheetnames:
        ws = wb.create_sheet("Inventario")
        _write_rows(ws, ["Sin datos"], [])

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
