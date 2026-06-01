from __future__ import annotations

import csv
import io
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

from utils.normalizer import clean_version


COLUMN_COUNT = 5


def _clean_text(raw: Any) -> str | None:
    if raw is None:
        return None
    val = str(raw).strip()
    if val in ("", "-", "–", "nan", "NaN", "None"):
        return None
    return val


def _parse_date(raw: Any):
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if hasattr(raw, "date") and not isinstance(raw, str):
        try:
            return raw.date()
        except Exception:
            pass
    val = str(raw).strip()
    if not val or val in ("-", "–", "nan", "NaN"):
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(val, fmt).date()
        except ValueError:
            continue
    return None


def _is_header(row: list[Any]) -> bool:
    values = [str(cell or "").strip().lower() for cell in row]
    joined = " ".join(values)
    if "nombre" in joined and ("editor" in joined or "fabricante" in joined or "proveedor" in joined):
        return True
    return bool(values and values[0] in {"nombre", "nombre del software"})


def _row_to_program(row: list[Any]) -> dict | None:
    row = list(row[:COLUMN_COUNT]) + [None] * max(0, COLUMN_COUNT - len(row))
    nombre = _clean_text(row[0])
    if not nombre:
        return None
    return {
        "nombre": nombre,
        "fabricante": _clean_text(row[1]),
        "fecha_instalacion": _parse_date(row[2]),
        "tamano": _clean_text(row[3]),
        "version": clean_version(row[4]),
    }


def _rows_to_programs(rows: list[list[Any]]) -> list[dict]:
    usable = [row for row in rows if any(_clean_text(cell) for cell in row)]
    if usable and _is_header(usable[0]):
        usable = usable[1:]
    programs = []
    for row in usable:
        item = _row_to_program(row)
        if item:
            programs.append(item)
    return programs


def parse_paste(text: str) -> list[dict]:
    rows: list[list[str]] = []
    for line in (text or "").splitlines():
        if not line.strip():
            continue
        rows.append(line.rstrip("\n").split("\t"))
    return _rows_to_programs(rows)


def _parse_csv(file_bytes: bytes) -> list[dict]:
    text = file_bytes.decode("utf-8-sig", errors="replace")
    best_rows: list[list[str]] = []
    for delimiter in ("\t", ";", ","):
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        rows = [row for row in reader]
        score = sum(1 for row in rows if len(row) >= COLUMN_COUNT)
        if score > sum(1 for row in best_rows if len(row) >= COLUMN_COUNT):
            best_rows = rows
    return _rows_to_programs(best_rows)


def _parse_xlsx(file_bytes: bytes) -> list[dict]:
    workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    sheet = workbook.worksheets[0]
    rows = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
    return _rows_to_programs(rows)


def parse_file(file_bytes: bytes, filename: str) -> list[dict]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        return _parse_csv(file_bytes)
    if suffix == ".xlsx":
        return _parse_xlsx(file_bytes)
    raise ValueError("Formato no soportado. Use .csv o .xlsx")


def dataframe_from_programs(programs: list[dict]) -> pd.DataFrame:
    return pd.DataFrame(programs)


EQUIPO_COLUMN_MAP = {
    "nombre del equipo": "nombre",
    "equipo": "nombre",
    "dispositivo": "nombre",
    "nombre": "nombre",
    "tipo de dispositivo": "tipo_dispositivo",
    "tipo dispositivo": "tipo_dispositivo",
    "tipo": "tipo_dispositivo",
    "sistema operativo": "sistema_operativo",
    "so": "sistema_operativo",
    "os": "sistema_operativo",
    "procesador": "procesador",
    "memoria ram": "ram",
    "ram": "ram",
    "memoria": "ram",
    "almacenamiento": "almacenamiento",
    "disco": "almacenamiento",
    "storage": "almacenamiento",
    "departamento": "departamento_nombre",
    "ubicación física": "ubicacion",
    "ubicacion fisica": "ubicacion",
    "ubicación": "ubicacion",
    "ubicacion": "ubicacion",
    "estado": "estado",
    "observaciones": "notas",
    "coste": "coste",
    "fecha de adquisición": "fecha_adquisicion",
    "fecha adquisicion": "fecha_adquisicion",
    "fecha_adquisicion": "fecha_adquisicion",
    "marca / modelo": "marca_modelo",
    "marca/modelo": "marca_modelo",
    "marca": "marca_modelo",
    "modelo": "marca_modelo",
    "nº de serie": "num_serie",
    "nº de serie bios": "num_serie",
    "nº serie": "num_serie",
    "numero de serie": "num_serie",
    "número de serie": "num_serie",
    "num serie": "num_serie",
    "serie": "num_serie",
    "dirección mac": "mac_address",
    "direccion mac": "mac_address",
    "dirección": "mac_address",
    "direccion": "mac_address",
    "mac": "mac_address",
    "mac address": "mac_address",
    "responsable": "responsable",
    "nº": None,
    "num": None,
    "numero": None,
    "número": None,
}


def _normalize_header(col: str) -> str:
    return col.strip().lower().replace("_", " ").replace("-", " ")


def _detect_equipo_headers(headers: list[str]) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for i, raw in enumerate(headers):
        norm = _normalize_header(raw)
        db_field = EQUIPO_COLUMN_MAP.get(norm)
        if db_field is not None:
            mapping[i] = db_field
    return mapping


def _is_equipo_header(row: list[Any]) -> bool:
    values = [_normalize_header(str(c or "")) for c in row]
    joined = " ".join(values)
    if "nombre del equipo" in joined or "nombre" in joined and "tipo" in joined:
        return True
    return False


def _parse_equipo_value(db_field: str, raw: Any) -> Any:
    text = _clean_text(raw)
    if text is None:
        return None
    if db_field == "coste":
        try:
            cleaned = text.replace(",", ".").replace(" ", "")
            return float(cleaned)
        except (ValueError, TypeError):
            return None
    if db_field == "fecha_adquisicion":
        return _parse_date(text)
    if db_field == "estado":
        norm = text.strip().lower()
        if norm in ("activo", "si", "sí", "true", "1"):
            return True
        if norm in ("inactivo", "no", "false", "0"):
            return False
        return None
    return text


def _row_to_equipo(headers_map: dict[int, str], row: list[Any]) -> dict | None:
    equipo: dict[str, Any] = {}
    has_name = False
    for i, raw in enumerate(row):
        if i not in headers_map:
            continue
        db_field = headers_map[i]
        value = _parse_equipo_value(db_field, raw)
        equipo[db_field] = value
        if db_field == "nombre" and value:
            has_name = True
    if not has_name:
        return None
    return equipo


def _rows_to_equipos(rows: list[list[Any]]) -> list[dict]:
    usable = [row for row in rows if any(_clean_text(cell) for cell in row)]
    if not usable:
        return []
    headers_map = {}
    if _is_equipo_header(usable[0]):
        headers_map = _detect_equipo_headers(usable[0])
        if not headers_map:
            return []
        usable = usable[1:]
    equipos = []
    for row in usable:
        item = _row_to_equipo(headers_map, row)
        if item:
            equipos.append(item)
    return equipos


def parse_equipos_paste(text: str) -> list[dict]:
    rows: list[list[str]] = []
    for line in (text or "").splitlines():
        if not line.strip():
            continue
        rows.append(line.rstrip("\n").split("\t"))
    return _rows_to_equipos(rows)


def _parse_equipos_csv(file_bytes: bytes) -> list[dict]:
    text = file_bytes.decode("utf-8-sig", errors="replace")
    best_rows: list[list[str]] = []
    for delimiter in ("\t", ";", ","):
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        rows = [row for row in reader]
        if len(rows) > len(best_rows):
            best_rows = rows
    return _rows_to_equipos(best_rows)


def _parse_equipos_xlsx(file_bytes: bytes) -> list[dict]:
    workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    sheet = workbook.worksheets[0]
    rows = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
    return _rows_to_equipos(rows)


def parse_equipos_file(file_bytes: bytes, filename: str) -> list[dict]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        return _parse_equipos_csv(file_bytes)
    if suffix == ".xlsx":
        return _parse_equipos_xlsx(file_bytes)
    raise ValueError("Formato no soportado. Use .csv o .xlsx")
